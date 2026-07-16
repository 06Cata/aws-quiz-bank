"""Export local CLF and SAA question JSON files to a two-sheet Excel workbook.

Usage:
    python scripts/export_questions_to_excel.py
    python scripts/export_questions_to_excel.py --output scripts/aws_quiz_bank.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - depends on the local environment
    raise SystemExit(
        "缺少 openpyxl，請先執行：python -m pip install openpyxl"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_QUESTIONS_DIR = ROOT / "questions"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "aws_quiz_bank.xlsx"
FILE_PATTERN = re.compile(
    r"^(?P<prefix>clf|saa)_Q(?P<start>\d+)-Q(?P<end>\d+)\.json$",
    re.IGNORECASE,
)
OPTION_KEYS = tuple("ABCDEF")
MAX_CELL_LENGTH = 32_767


@dataclass(frozen=True)
class ExamConfig:
    prefix: str
    sheet_name: str
    expected_exam: str


EXAMS = (
    ExamConfig("clf", "CLF-C02", "AWS CLF-C02"),
    ExamConfig("saa", "SAA-C03", "AWS SAA-C03"),
)


BASE_HEADERS = (
    "source_file",
    "question_no",
    "exam",
    "domain",
    "selection_type",
    "correct_answers",
    "question_zh",
    "question_en",
)

TAIL_HEADERS = (
    "answer_zh",
    "answer_en",
    "discussion_zh",
    "discussion_en",
)


def option_headers() -> list[str]:
    headers: list[str] = []
    for key in OPTION_KEYS:
        headers.extend(
            (
                f"option_{key}_zh",
                f"option_{key}_en",
                f"explanation_{key}_zh",
                f"explanation_{key}_en",
            )
        )
    return headers


HEADERS = [*BASE_HEADERS, *option_headers(), *TAIL_HEADERS]


def required_text(value: Any, location: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{location} 必須是非空字串")
    text = ILLEGAL_CHARACTERS_RE.sub("", value.strip())
    if len(text) > MAX_CELL_LENGTH:
        raise ValueError(f"{location} 超過 Excel 儲存格 {MAX_CELL_LENGTH} 字元限制")
    return text


def bilingual(value: Any, location: str) -> tuple[str, str]:
    if not isinstance(value, dict):
        raise ValueError(f"{location} 必須是包含 zh、en 的物件")
    return (
        required_text(value.get("zh"), f"{location}.zh"),
        required_text(value.get("en"), f"{location}.en"),
    )


def discover_files(directory: Path, prefix: str) -> list[tuple[int, int, Path]]:
    files: list[tuple[int, int, Path]] = []
    for path in directory.glob(f"{prefix}_Q*-Q*.json"):
        match = FILE_PATTERN.fullmatch(path.name)
        if not match or match.group("prefix").lower() != prefix:
            continue
        start = int(match.group("start"))
        end = int(match.group("end"))
        if start > end:
            raise ValueError(f"{path.name} 起始題號不可大於結束題號")
        files.append((start, end, path))
    return sorted(files, key=lambda item: (item[0], item[1], item[2].name.lower()))


def question_to_row(question: Any, source_file: str, exam: str, location: str) -> list[Any]:
    if not isinstance(question, dict):
        raise ValueError(f"{location} 必須是物件")

    question_no = question.get("question_no")
    if not isinstance(question_no, int) or isinstance(question_no, bool) or question_no < 1:
        raise ValueError(f"{location}.question_no 必須是正整數")

    question_zh, question_en = bilingual(question.get("question_text"), f"{location}.question_text")
    options = question.get("options")
    explanations = question.get("option_explanations")
    if not isinstance(options, dict) or not options:
        raise ValueError(f"{location}.options 必須是非空物件")
    if not isinstance(explanations, dict) or options.keys() != explanations.keys():
        raise ValueError(f"{location} 的 options 與 option_explanations 代號不一致")

    correct_answers = question.get("correct_answers")
    if not isinstance(correct_answers, list) or not correct_answers:
        raise ValueError(f"{location}.correct_answers 必須是非空陣列")
    correct_text = ",".join(required_text(str(key), f"{location}.correct_answers") for key in correct_answers)

    row: list[Any] = [
        source_file,
        question_no,
        exam,
        required_text(question.get("domain"), f"{location}.domain"),
        required_text(question.get("selection_type"), f"{location}.selection_type"),
        correct_text,
        question_zh,
        question_en,
    ]

    for key in OPTION_KEYS:
        if key in options:
            option_zh, option_en = bilingual(options[key], f"{location}.options.{key}")
            explanation_zh, explanation_en = bilingual(
                explanations[key], f"{location}.option_explanations.{key}"
            )
            row.extend((option_zh, option_en, explanation_zh, explanation_en))
        else:
            row.extend(("", "", "", ""))

    answer_zh, answer_en = bilingual(question.get("answer_text"), f"{location}.answer_text")
    discussion_zh, discussion_en = bilingual(question.get("discussion"), f"{location}.discussion")
    row.extend((answer_zh, answer_en, discussion_zh, discussion_en))
    return row


def load_exam_rows(directory: Path, config: ExamConfig) -> tuple[list[list[Any]], int]:
    files = discover_files(directory, config.prefix)
    if not files:
        raise FileNotFoundError(f"找不到 {config.prefix}_Q*-Q*.json")

    rows: list[list[Any]] = []
    numbers: list[int] = []
    for file_start, file_end, path in files:
        with path.open(encoding="utf-8-sig") as file:
            document = json.load(file)
        if not isinstance(document, dict) or not isinstance(document.get("questions"), list):
            raise ValueError(f"{path.name} 必須是包含 questions 陣列的 JSON 物件")
        exam = required_text(document.get("exam"), f"{path.name}.exam")
        if exam != config.expected_exam:
            raise ValueError(
                f"{path.name}.exam 應為 {config.expected_exam}，目前為 {exam}"
            )

        questions = document["questions"]
        actual_numbers = [question.get("question_no") for question in questions]
        expected_numbers = list(range(file_start, file_end + 1))
        if actual_numbers != expected_numbers:
            raise ValueError(
                f"{path.name} 題號應為 Q{file_start}-Q{file_end}，實際為 {actual_numbers}"
            )

        for index, question in enumerate(questions):
            rows.append(
                question_to_row(
                    question,
                    path.name,
                    exam,
                    f"{path.name}.questions[{index}]",
                )
            )
        numbers.extend(actual_numbers)

    expected_all = list(range(1, numbers[-1] + 1))
    if numbers != expected_all:
        raise ValueError(
            f"{config.prefix.upper()} 題號必須從 Q1 開始連續排列；目前題號不連續"
        )
    return rows, len(files)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30

    compact_columns = {
        "source_file": 24,
        "question_no": 12,
        "exam": 16,
        "domain": 34,
        "selection_type": 14,
        "correct_answers": 16,
    }
    for index, header in enumerate(HEADERS, start=1):
        width = compact_columns.get(header, 42 if header.endswith(("_zh", "_en")) else 20)
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(questions_dir: Path, output: Path) -> dict[str, dict[str, int]]:
    if not questions_dir.is_dir():
        raise FileNotFoundError(f"找不到 questions 資料夾：{questions_dir}")

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary: dict[str, dict[str, int]] = {}

    for config in EXAMS:
        rows, file_count = load_exam_rows(questions_dir, config)
        sheet = workbook.create_sheet(config.sheet_name)
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(row)
        style_sheet(sheet)
        summary[config.sheet_name] = {"files": file_count, "questions": len(rows)}

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.stem}.tmp{output.suffix}")
    workbook.save(temporary)
    temporary.replace(output)

    # Reopen once to ensure the generated workbook is structurally readable.
    verified = load_workbook(output, read_only=True)
    if verified.sheetnames != [config.sheet_name for config in EXAMS]:
        raise RuntimeError(f"Excel sheet 驗證失敗：{verified.sheetnames}")
    verified.close()
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="將 questions 內的 CLF 與 SAA JSON 匯出為兩個 Excel sheets"
    )
    parser.add_argument(
        "--questions-dir",
        type=Path,
        default=DEFAULT_QUESTIONS_DIR,
        help=f"JSON 題庫目錄（預設：{DEFAULT_QUESTIONS_DIR}）",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"輸出 Excel 路徑（預設：{DEFAULT_OUTPUT}）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    questions_dir = args.questions_dir.expanduser().resolve()
    output = args.output.expanduser().resolve()
    summary = build_workbook(questions_dir, output)
    for sheet_name, counts in summary.items():
        print(
            f"{sheet_name}: files={counts['files']}, questions={counts['questions']}"
        )
    print(f"Excel written: {output}")


if __name__ == "__main__":
    main()
